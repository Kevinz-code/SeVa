import json
from openpyxl import Workbook


def eval_all(ans_file, label_file):
    answers = [json.loads(q) for q in open(ans_file, 'r')]
    label_list = [json.loads(q)['label'] for q in open(label_file, 'r')]

    for answer in answers:
        text = answer['answer']

        # Only keep the first sentence
        if text.find('.') != -1:
            text = text.split('.')[0]

        text = text.replace(',', '')
        words = text.split(' ')
        if 'No' in words or 'not' in words or 'no' in words:
            answer['answer'] = 'no'
        else:
            answer['answer'] = 'yes'

    for i in range(len(label_list)):
        if label_list[i] == 'no':
            label_list[i] = 0
        else:
            label_list[i] = 1

    pred_list = []
    for answer in answers:
        if answer['answer'] == 'no':
            pred_list.append(0)
        else:
            pred_list.append(1)

    pos = 1
    neg = 0
    yes_ratio = pred_list.count(1) / len(pred_list)

    TP, TN, FP, FN = 0, 0, 0, 0
    for pred, label in zip(pred_list, label_list):
        if pred == pos and label == pos:
            TP += 1
        elif pred == pos and label == neg:
            FP += 1
        elif pred == neg and label == neg:
            TN += 1
        elif pred == neg and label == pos:
            FN += 1

    print('TP\tFP\tTN\tFN\t')
    print('{}\t{}\t{}\t{}'.format(TP, FP, TN, FN))

    precision = float(TP) / float(TP + FP)
    recall = float(TP) / float(TP + FN)
    f1 = 2*precision*recall / (precision + recall)
    acc = (TP + TN) / (TP + TN + FP + FN)
    print('Accuracy: {}'.format(acc))
    print('Precision: {}'.format(precision))
    print('Recall: {}'.format(recall))
    print('F1 score: {}'.format(f1))
    print('Yes ratio: {}'.format(yes_ratio))


    acc_percentage = "{:.2f}".format(acc *100)
    precision_percentage = "{:.2f}".format(precision * 100)
    f1_percentage = "{:.2f}".format(f1 * 100)
    yes_ratio_percentage = "{:.2f}".format(yes_ratio * 100)

    return acc_percentage, precision_percentage, f1_percentage, yes_ratio_percentage

import argparse

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--path", default="")
    args = parser.parse_args()

    ans_file_list = ['{}/pope_random.jsonl'.format(args.path),
                     '{}/pope_popular.jsonl'.format(args.path),
                     '{}/pope_adv.jsonl'.format(args.path)]
    label_file_list = ['/data/hypertext/zhuk/HA-DPO/ha_dpo/data/POPE/output/coco/coco_pope_adversarial.json',
                  '/data/hypertext/zhuk/HA-DPO/ha_dpo/data/POPE/output/coco/coco_pope_adversarial.json',
                  '/data/hypertext/zhuk/HA-DPO/ha_dpo/data/POPE/output/coco/coco_pope_adversarial.json']
    tag_list = ['random', 'popular', 'adversarial']

    # 创建一个新的Workbook对象
    wb = Workbook()

    # 选择活动的工作表
    ws = wb.active

    # 在工作表中写入表头
    ws['A1'] = 'Tag'
    ws['B1'] = 'Accuracy'
    ws['C1'] = 'Precision'
    ws['D1'] = 'F1 Score'
    ws['E1'] = 'Yes Ratio'



    for i in range(3):
        print(tag_list[i])
        print("------------------------------------------")
        acc, precision, f1, yes_ratio = eval_all(ans_file_list[i], label_file_list[i])
        print("------------------------------------------")
        print()

        # 将数据写入Excel表格
        row = i + 2  # 行数从第2行开始
        ws.cell(row=row, column=1, value=tag_list[i])
        ws.cell(row=row, column=2, value=acc)
        ws.cell(row=row, column=3, value=precision)
        ws.cell(row=row, column=4, value=f1)
        ws.cell(row=row, column=5, value=yes_ratio)

    # 保存工作簿到文件
    wb.save('{}/evaluation_results.xlsx'.format(args.path))